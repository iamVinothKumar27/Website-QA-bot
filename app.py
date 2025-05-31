import os
import json
import requests
from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_session import Session
from werkzeug.utils import secure_filename
from uuid import uuid4
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from authlib.integrations.flask_client import OAuth
from PyPDF2 import PdfReader
import docx
import openpyxl

from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings
import google.generativeai as genai
from langchain_community.vectorstores import FAISS
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate

load_dotenv()
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

app = Flask(__name__)
app.secret_key = "secret"
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
USERS_FILE = "users.json"
ADMIN_CREDENTIALS = {"email": "admin@site.com", "password": "admin123"}

# Auth0 setup
oauth = OAuth(app)
auth0 = oauth.register(
    'auth0',
    client_id=os.getenv("AUTH0_CLIENT_ID"),
    client_secret=os.getenv("AUTH0_CLIENT_SECRET"),
    api_base_url=f'https://{os.getenv("AUTH0_DOMAIN")}',
    access_token_url=f'https://{os.getenv("AUTH0_DOMAIN")}/oauth/token',
    authorize_url=f'https://{os.getenv("AUTH0_DOMAIN")}/authorize',
    client_kwargs={
        'scope': 'openid profile email',
    },
)

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r") as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=2)

@app.before_request
def require_login():
    allowed_routes = ['login', 'signup', 'auth0_login', 'callback', 'static']
    if request.endpoint not in allowed_routes and 'user' not in session and 'admin' not in session:
        return redirect(url_for('login'))

@app.context_processor
def inject_user_initials():
    if 'user' in session:
        users = load_users()
        user = users.get(session['user'], {})
        name = user.get("username", "")
        initials = ''.join(part[0].upper() for part in name.split() if part)
        return dict(user_initials=initials, user_name=name)
    elif 'profile' in session:
        name = session['profile'].get('name', '')
        initials = ''.join(part[0].upper() for part in name.split() if part)
        return dict(user_initials=initials, user_name=name)
    return dict(user_initials="", user_name="")

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form["username"]
        email = request.form["email"]
        password = request.form["password"]
        users = load_users()
        if email in users:
            flash("Email already registered!", "info")
            return redirect(url_for("login"))
        else:
            users[email] = {"username": username, "password": password, "chats": {}}
            save_users(users)
            flash("Signup successful. Please login.", "success")
            return redirect(url_for("login"))
    return render_template("signup.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]
        users = load_users()
        if email == ADMIN_CREDENTIALS["email"] and password == ADMIN_CREDENTIALS["password"]:
            session.clear()
            session["admin"] = True
            return redirect(url_for("admin_panel"))
        user = users.get(email)
        if user and user["password"] == password:
            session.clear()
            session["user"] = email
            session["chats"] = user.get("chats", {})
            return redirect(url_for("chat"))
        flash("Invalid credentials", "danger")
    return render_template("login.html")

@app.route("/auth0-login")
def auth0_login():
    return auth0.authorize_redirect(redirect_uri=os.getenv("AUTH0_CALLBACK_URL"))

@app.route("/callback")
def callback():
    token = oauth.auth0.authorize_access_token()
    userinfo = oauth.auth0.get("userinfo").json()  # ✅ Correct way
    session["profile"] = {
        "user_id": userinfo["sub"],
        "name": userinfo["name"],
        "email": userinfo["email"]
    }
    session["user"] = userinfo["email"]
    session["chats"] = load_users().get(userinfo["email"], {}).get("chats", {})
    return redirect(url_for("chat"))


@app.route("/auth0-logout")
def auth0_logout():
    session.clear()
    return redirect(
        f"https://{os.getenv('AUTH0_DOMAIN')}/v2/logout?returnTo={url_for('login', _external=True)}&client_id={os.getenv('AUTH0_CLIENT_ID')}"
    )

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/admin")
def admin_panel():
    if not session.get("admin"):
        return redirect(url_for("login"))
    users = load_users()
    return render_template("admin.html", users=users)

@app.route("/admin/delete_user/<email>")
def delete_user(email):
    if not session.get("admin"):
        return redirect(url_for("login"))
    users = load_users()
    if email in users:
        del users[email]
        save_users(users)
    return redirect(url_for("admin_panel"))

@app.route("/")
def index():
    return redirect(url_for("chat"))

@app.route("/chat", methods=["GET", "POST"])
def chat():
    if "chats" not in session:
        session["chats"] = {}

    chat_id = request.args.get("chat_id") or session.get("current_chat")

    if chat_id:
        chat = session["chats"].get(chat_id, {})
    else:
        chat = {}

    chat.setdefault("history", [])
    chat.setdefault("title", "Untitled Chat")
    chat.setdefault("url", None)
    chat.setdefault("uploaded_docs", [])
    chat.setdefault("vectorized", False)

    url = chat["url"]
    uploaded_docs = chat["uploaded_docs"]
    question = ""
    answer = ""
    status = ""

    if request.method == "POST":
        question = request.form.get("question")

        if not chat.get("vectorized", False):
            url = request.form.get("url")
            uploaded_files = request.files.getlist("documents")
            combined_text = ""

            if url:
                extracted_text, soup = extract_text_from_url(url)
                combined_text += extracted_text
                chat["url"] = url
                try:
                    site_title = soup.title.string.strip()
                    if site_title:
                        chat["title"] = site_title[:50]
                except Exception:
                    pass

            if uploaded_files and uploaded_files[0].filename:
                file_text, filenames = extract_text_from_files(uploaded_files)
                combined_text += file_text
                chat["uploaded_docs"] = filenames
                chat["title"] = filenames[0] if filenames else chat["title"]

            if combined_text:
                chunks = get_text_chunks(combined_text)
                get_vector_store(chunks)
                chat["vectorized"] = True
                chat_id = str(uuid4())
                session["current_chat"] = chat_id
                session["chats"][chat_id] = chat
                status = "complete"
            else:
                answer = "Failed to extract content from the provided inputs."
                status = "error"
        elif question:
            answer = get_answer(question)
            chat["history"].append({"question": question, "answer": answer})
            status = "complete"

        session.modified = True

    return render_template(
        "index.html",
        question=question,
        status=status,
        chat_history=chat["history"],
        chat_url=chat["url"],
        uploaded_docs=chat["uploaded_docs"],
        chat_titles={cid: c["title"] for cid, c in session["chats"].items()}
    )

@app.route("/rename", methods=["POST"])
def rename_chat():
    chat_id = request.form.get("chat_id")
    new_title = request.form.get("new_title")
    if chat_id in session.get("chats", {}):
        session["chats"][chat_id]["title"] = new_title
        session.modified = True
    return redirect("/chat")

@app.route("/delete", methods=["POST"])
def delete_chat():
    chat_id = request.form.get("chat_id")
    if chat_id in session.get("chats", {}):
        session["chats"].pop(chat_id)
        if session.get("current_chat") == chat_id:
            session["current_chat"] = next(iter(session["chats"]), None)
        session.modified = True
    return redirect("/chat")

def extract_text_from_url(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.content, "html.parser")
        for script in soup(["script", "style"]):
            script.decompose()
        return soup.get_text(separator="\n"), soup
    except Exception as e:
        return f"Failed to fetch content: {e}", None

def extract_text_from_files(files):
    full_text = ""
    filenames = []
    for file in files:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        filenames.append(filename)

        if filename.endswith(".pdf"):
            reader = PdfReader(file_path)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    full_text += text
        elif filename.endswith(".docx"):
            doc = docx.Document(file_path)
            for para in doc.paragraphs:
                full_text += para.text + "\n"
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            for sheet in wb:
                headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    full_text += '\n'.join(f"{h}: {c}" for h, c in zip(headers, row)) + '\n'
    return full_text, filenames

def get_text_chunks(text):
    splitter = RecursiveCharacterTextSplitter(chunk_size=10000, chunk_overlap=1000)
    return splitter.split_text(text)

def get_vector_store(text_chunks):
    embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    vector_store = FAISS.from_texts(text_chunks, embedding=embeddings)
    vector_store.save_local("vectors_storage")

def get_conversational_chain():
    prompt_template = """
    You are a helpful assistant that reads the context and answers the question.
    Be detailed. If unsure, say "I couldn't find that in the website or documents."

    Context:
    {context}

    Question:
    {question}

    Answer:
    """
    model = ChatGoogleGenerativeAI(model="models/gemini-1.5-flash", temperature=0.3)
    prompt = PromptTemplate(template=prompt_template, input_variables=["context", "question"])
    return load_qa_chain(model, chain_type="stuff", prompt=prompt)

def get_answer(query):
    embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    vector_store = FAISS.load_local("vectors_storage", embeddings, allow_dangerous_deserialization=True)
    docs = vector_store.similarity_search(query)
    chain = get_conversational_chain()
    result = chain.invoke({"input_documents": docs, "question": query})
    return result["output_text"]

if __name__ == "__main__":
    app.run(debug=True)